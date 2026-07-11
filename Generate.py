from __future__ import annotations

import html
import re
import sys
from pathlib import Path

import qrcode
from openpyxl import load_workbook
from PIL import Image, ImageDraw
from qrcode.constants import ERROR_CORRECT_H


# ============================================================
# PROJECT SETTINGS
# ============================================================

PROJECT_FOLDER = Path(__file__).resolve().parent

SPREADSHEET_FILE = PROJECT_FOLDER / "Coin Spreadsheet.xlsx"
TEMPLATE_FILE = PROJECT_FOLDER / "template.html"

IMAGES_FOLDER = PROJECT_FOLDER / "images"
PAGES_FOLDER = PROJECT_FOLDER / "pages"
ICONS_FOLDER = PROJECT_FOLDER / "icons"
QR_FOLDER = PROJECT_FOLDER / "qr"

SITE_ROOT = "https://khanstruct.github.io/Remembrance-Coins"
PAGES_URL = f"{SITE_ROOT}/pages"


# ============================================================
# COLOR NAMES
# ============================================================

COLORS = {
    "amber": "rgba(255, 180, 80, 0.45)",
    "black": "rgba(0, 0, 0, 0.00)",
    "blue": "rgba(90, 160, 255, 0.48)",
    "dark blue": "rgba(25, 55, 130, 0.48)",
    "dark gray": "rgba(65, 70, 80, 0.42)",
    "dark green": "rgba(25, 95, 55, 0.46)",
    "dark red": "rgba(125, 20, 30, 0.48)",
    "green": "rgba(75, 200, 110, 0.46)",
    "light blue": "rgba(160, 220, 255, 0.52)",
    "light gray": "rgba(205, 215, 225, 0.48)",
    "light green": "rgba(165, 235, 170, 0.48)",
    "orange": "rgba(255, 120, 30, 0.48)",
    "pink": "rgba(255, 130, 190, 0.48)",
    "purple": "rgba(155, 90, 255, 0.48)",
    "red": "rgba(255, 65, 55, 0.48)",
    "white": "rgba(255, 255, 255, 0.54)",
    "yellow": "rgba(255, 225, 90, 0.50)",
}


# ============================================================
# CATEGORY ICONS
# ============================================================

CATEGORY_ICONS = {
    "common": ICONS_FOLDER / "common.png",
    "meaningful": ICONS_FOLDER / "meaningful.png",
    "clue": ICONS_FOLDER / "clue.png",
}


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_value(value: object) -> str:
    """Convert an Excel cell value into clean text."""
    if value is None:
        return ""

    return str(value).strip()


def get_color(
    color_name: str,
    row_number: int,
    column_name: str,
) -> str:
    """Convert a named spreadsheet color into an RGBA value."""
    normalized_name = color_name.strip().lower()

    if normalized_name not in COLORS:
        available = ", ".join(
            sorted(name.title() for name in COLORS)
        )

        raise ValueError(
            f'Unknown color "{color_name}" in row {row_number}, '
            f'column "{column_name}".\n'
            f"Available colors: {available}"
        )

    return COLORS[normalized_name]


def normalize_page_filename(page_name: str) -> str:
    """Ensure the generated webpage ends with .html."""
    page_name = page_name.strip()

    if not page_name.lower().endswith(".html"):
        page_name += ".html"

    return page_name


def resolve_image_filename(image_name: str) -> str:
    """
    Use the supplied image filename.

    If no extension is present, search for common image formats.
    """
    image_name = image_name.strip()
    supplied_path = Path(image_name)

    if supplied_path.suffix:
        return image_name

    possible_extensions = [
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
    ]

    for extension in possible_extensions:
        candidate = IMAGES_FOLDER / f"{image_name}{extension}"

        if candidate.exists():
            return candidate.name

    return f"{image_name}.png"


def text_to_paragraphs(memory_text: str) -> str:
    """
    Convert Excel line breaks into separate HTML paragraphs.

    One or more line breaks begin a new paragraph.
    """
    normalized = (
        memory_text
        .replace("\r\n", "\n")
        .replace("\r", "\n")
    )

    paragraphs = [
        paragraph.strip()
        for paragraph in re.split(r"\n+", normalized)
        if paragraph.strip()
    ]

    return "\n        ".join(
        f"<p>{html.escape(paragraph)}</p>"
        for paragraph in paragraphs
    )


def validate_required_columns(headers: dict[str, int]) -> None:
    """Confirm that the spreadsheet includes every required header."""
    required_columns = {
        "page",
        "title",
        "image",
        "glow start",
        "glow end",
        "text",
        "category",
    }

    missing = required_columns - set(headers)

    if missing:
        missing_list = ", ".join(
            sorted(column.title() for column in missing)
        )

        raise ValueError(
            f"Spreadsheet is missing required column(s): "
            f"{missing_list}"
        )


def get_category_icon(
    category: str,
    row_number: int,
) -> Path:
    """Return the correct center icon for a QR code."""
    normalized_category = category.strip().lower()

    if normalized_category not in CATEGORY_ICONS:
        available = ", ".join(
            name.title() for name in CATEGORY_ICONS
        )

        raise ValueError(
            f'Unknown category "{category}" in row {row_number}.\n'
            f"Available categories: {available}"
        )

    icon_path = CATEGORY_ICONS[normalized_category]

    if not icon_path.exists():
        raise FileNotFoundError(
            f'Category icon not found for "{category}":\n'
            f"{icon_path}"
        )

    return icon_path


# ============================================================
# QR CODE GENERATION
# ============================================================

def generate_qr_code(
    url: str,
    icon_path: Path,
    output_path: Path,
) -> None:
    """
    Generate a high-resolution QR code and place the category icon
    in its center.

    Error-correction level H is used because the center icon covers
    part of the QR pattern.
    """
    qr = qrcode.QRCode(
        version=None,
        error_correction=ERROR_CORRECT_H,
        box_size=20,
        border=4,
    )

    qr.add_data(url)
    qr.make(fit=True)

    qr_image = qr.make_image(
        fill_color="black",
        back_color="white",
    ).convert("RGBA")

    icon = Image.open(icon_path).convert("RGBA")

    # Keep the icon modestly sized so scanning remains reliable.
    icon_size = int(qr_image.width * 0.18)

    icon.thumbnail(
        (icon_size, icon_size),
        Image.Resampling.LANCZOS,
    )

    # Create a white backing plate around the icon.
    padding = max(8, int(qr_image.width * 0.015))

    badge_width = icon.width + padding * 2
    badge_height = icon.height + padding * 2

    badge = Image.new(
        "RGBA",
        (badge_width, badge_height),
        (255, 255, 255, 255),
    )

    badge_draw = ImageDraw.Draw(badge)

    corner_radius = max(8, int(min(badge.size) * 0.12))

    badge_draw.rounded_rectangle(
        (0, 0, badge_width - 1, badge_height - 1),
        radius=corner_radius,
        fill=(255, 255, 255, 255),
    )

    icon_x = (badge_width - icon.width) // 2
    icon_y = (badge_height - icon.height) // 2

    badge.alpha_composite(
        icon,
        (icon_x, icon_y),
    )

    badge_x = (qr_image.width - badge_width) // 2
    badge_y = (qr_image.height - badge_height) // 2

    qr_image.alpha_composite(
        badge,
        (badge_x, badge_y),
    )

    qr_image.convert("RGB").save(
        output_path,
        format="PNG",
        optimize=True,
    )


# ============================================================
# MAIN GENERATOR
# ============================================================

def generate_project() -> None:
    """Generate all webpages and matching QR codes."""
    if not SPREADSHEET_FILE.exists():
        raise FileNotFoundError(
            f"Spreadsheet not found:\n{SPREADSHEET_FILE}"
        )

    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(
            f"Template not found:\n{TEMPLATE_FILE}"
        )

    if not IMAGES_FOLDER.exists():
        raise FileNotFoundError(
            f"Images folder not found:\n{IMAGES_FOLDER}"
        )

    if not ICONS_FOLDER.exists():
        raise FileNotFoundError(
            f"Icons folder not found:\n{ICONS_FOLDER}"
        )

    PAGES_FOLDER.mkdir(exist_ok=True)
    QR_FOLDER.mkdir(exist_ok=True)

    template = TEMPLATE_FILE.read_text(
        encoding="utf-8",
    )

    workbook = load_workbook(
        SPREADSHEET_FILE,
        read_only=True,
        data_only=True,
    )

    worksheet = workbook.active

    header_values = next(
        worksheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True,
        )
    )

    headers = {
        clean_value(header).lower(): index
        for index, header in enumerate(header_values)
        if clean_value(header)
    }

    validate_required_columns(headers)

    generated_pages = 0
    generated_qr_codes = 0
    warnings: list[str] = []
    seen_page_names: set[str] = set()

    category_counts = {
        "common": 0,
        "meaningful": 0,
        "clue": 0,
    }

    print("\nGenerating remembrance archive...\n")

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        ),
        start=2,
    ):
        page = clean_value(row[headers["page"]])
        title = clean_value(row[headers["title"]])
        image = clean_value(row[headers["image"]])
        glow_start_name = clean_value(
            row[headers["glow start"]]
        )
        glow_end_name = clean_value(
            row[headers["glow end"]]
        )
        memory_text = clean_value(row[headers["text"]])
        category = clean_value(row[headers["category"]])

        values = [
            page,
            title,
            image,
            glow_start_name,
            glow_end_name,
            memory_text,
            category,
        ]

        # Ignore completely blank rows.
        if not any(values):
            continue

        missing_fields = []

        if not page:
            missing_fields.append("Page")
        if not title:
            missing_fields.append("Title")
        if not image:
            missing_fields.append("Image")
        if not glow_start_name:
            missing_fields.append("Glow Start")
        if not glow_end_name:
            missing_fields.append("Glow End")
        if not memory_text:
            missing_fields.append("Text")
        if not category:
            missing_fields.append("Category")

        if missing_fields:
            warnings.append(
                f"Row {row_number}: missing "
                f"{', '.join(missing_fields)}"
            )
            continue

        page_filename = normalize_page_filename(page)
        normalized_page_name = page_filename.lower()

        if normalized_page_name in seen_page_names:
            warnings.append(
                f'Row {row_number}: duplicate page filename '
                f'"{page_filename}"'
            )
            continue

        seen_page_names.add(normalized_page_name)

        image_filename = resolve_image_filename(image)
        image_path = IMAGES_FOLDER / image_filename

        if not image_path.exists():
            warnings.append(
                f'Row {row_number}: image not found: '
                f'"{image_filename}"'
            )

        glow_start = get_color(
            glow_start_name,
            row_number,
            "Glow Start",
        )

        glow_end = get_color(
            glow_end_name,
            row_number,
            "Glow End",
        )

        icon_path = get_category_icon(
            category,
            row_number,
        )

        normalized_category = category.lower()
        category_counts[normalized_category] += 1

        paragraphs = text_to_paragraphs(memory_text)

        replacements = {
            "{{TITLE}}": html.escape(title),
            "{{IMAGE}}": html.escape(
                image_filename,
                quote=True,
            ),
            "{{ALT_TEXT}}": html.escape(
                title,
                quote=True,
            ),
            "{{GLOW_START}}": glow_start,
            "{{GLOW_END}}": glow_end,
            "{{TEXT}}": paragraphs,
        }

        page_html = template

        for placeholder, replacement in replacements.items():
            page_html = page_html.replace(
                placeholder,
                replacement,
            )

        page_output_path = PAGES_FOLDER / page_filename

        page_output_path.write_text(
            page_html,
            encoding="utf-8",
        )

        generated_pages += 1

        page_url = f"{PAGES_URL}/{page_filename}"

        qr_filename = (
            f"{Path(page_filename).stem}.png"
        )

        qr_output_path = QR_FOLDER / qr_filename

        generate_qr_code(
            url=page_url,
            icon_path=icon_path,
            output_path=qr_output_path,
        )

        generated_qr_codes += 1

        print(
            f"  Created: {page_filename} "
            f"+ {qr_filename}"
        )

    workbook.close()

    print("\nBuild complete.")
    print(f"  Webpages: {generated_pages}")
    print(f"  QR codes: {generated_qr_codes}")
    print()
    print(f"  Common: {category_counts['common']}")
    print(f"  Meaningful: {category_counts['meaningful']}")
    print(f"  Clue: {category_counts['clue']}")

    if warnings:
        print("\nWarnings:")

        for warning in warnings:
            print(f"  - {warning}")

    print(f"\nPages folder:\n{PAGES_FOLDER}")
    print(f"\nQR folder:\n{QR_FOLDER}\n")


def main() -> None:
    try:
        generate_project()

    except Exception as error:
        print("\nThe generator encountered an error:\n")
        print(error)
        print()
        input("Press Enter to close...")
        sys.exit(1)

    input("Press Enter to close...")


if __name__ == "__main__":
    main()